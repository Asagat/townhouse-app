"""update_counterparties_from_xlsx.py — наполнение справочника «Контрагенты»
(counterparties) из листа «Контрагенты» файла templates/Миграция данных FTH.xlsx.

Выполняется идемпотентно:
  - читает строки листа (Код контрагента, Код квартиры, Фамилия, Имя, Отчество);
  - приводит к единой записи справочника по совпадению полного ФИО (нормализация);
  - если записи нет — создаёт; существующие — не дублирует/ не перезаписывает ФИО,
    но при необходимости обновляет контактные поля (не указаны → без изменений);
  - выводит статистику и сверку «сколько совпало / создано», не трогая привязок
    квартир/операций (их перепривязка — в рамках планируемого history-перезалива).

Запуск от каталога backend с указанием целевой БД:
    DATABASE_URL=... python migrations/update_counterparties_from_xlsx.py
"""

import os
import re
import sys

from pathlib import Path

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from database import SessionLocal  # noqa: E402
from models import Counterparty  # noqa: E402

XLSX = Path(os.environ.get("MIGRATION_XLSX", "")) if os.environ.get("MIGRATION_XLSX") else (
    Path(__file__).resolve().parents[2] / "templates" / "Миграция данных FTH.xlsx"
)

def _norm(name: str) -> str:
    return re.sub(r"\s+", " ", (name or "").strip().lower())


def main() -> int:
    try:
        import openpyxl
    except Exception as e:  # noqa: BLE001
        print("Нужен openpyxl:", e)
        return 2
    if not XLSX.exists():
        print("Не найден файл:", XLSX)
        return 2

    wb = openpyxl.load_workbook(str(XLSX), data_only=True)
    ws = wb["Контрагенты"]
    hdr = [h.value for h in next(ws.iter_rows(min_row=1, max_row=1))]

    def col(name: str) -> int:
        return [i for i, h in enumerate(hdr) if h == name][0]

    ci_code, ci_apt = col("Код контрагента"), col("Код квартиры")
    ci_f, ci_i, ci_o = col("Фамилия"), col("Имя"), col("Отчество")

    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r is None or r[ci_code] is None:
            continue
        parts = [str(r[ci_f] or ""), str(r[ci_i] or ""), str(r[ci_o] or "")]
        parts = [p.strip() for p in parts if p.strip()]
        full = " ".join(parts)
        rows.append((r[ci_code], r[ci_apt], full))
    print("Строк в листе: ", len(rows))

    db = SessionLocal()
    try:
        existing = db.query(Counterparty).all()
        by_norm = {}
        for o in existing:
            candidates = [o.first_name or "", o.last_name or "", o.middle_name or ""]
            # попробуем собрать из доступных полей; на случай разнобоя используем full_name
            key = _norm(o.full_name)
            by_norm.setdefault(key, []).append(o)
        added = matched = 0
        report = []
        for code, apt, full in rows:
            key = _norm(full)
            obj = (by_norm.get(key) or [None])[0]
            if obj is None:
                # попытка матча иначе: last+first без отчества
                parts = full.split()
                if parts:
                    fk = _norm(" ".join([parts[0], parts[1] if len(parts) > 1 else ""]))
            else:
                matched += 1
                report.append((int(code), full, obj.id, "existing"))
                continue
            # создаём
            parts = full.split()
            first = parts[1] if len(parts) > 1 else ""
            last = parts[0]
            middle = parts[2] if len(parts) > 2 else ""
            o = Counterparty(
                full_name=full,
                first_name=first or full,
                last_name=last,
                middle_name=middle,
                is_active=True,
            )
            db.add(o)
            db.flush()
            added += 1
            report.append((int(code), full, o.id, "created"))
            # память последующих совпадений
            by_norm.setdefault(_norm(full), [o])  # но если уже было None по ключу
        db.commit()

        print(f"Итого: совпало существующих = {matched}, создано = {added}, всего записей = "
              f"{db.query(Counterparty).count()}")
        print("Пример строк:")
        for c, name, oid, how in report:
            print(f"  {c:>3} | {name[:32]:32s} -> id {oid} [{how}]")
    except Exception as exc:  # noqa: BLE001
        db.rollback()
        print("Ошибка:", exc)
        return 1
    finally:
        db.close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
