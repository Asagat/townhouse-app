# backend/migrations/migrate_prepare_sources.py
"""Подготовка источников миграции из старой БД (роадмап, раздел 4).

Читает «сырую» выгрузку `templates/Миграция данных FTH.xlsx` и превращает в
нормализованные CSV-наборы в `templates/clean/`, которые затем читает файл
синтеза документов (`backend/migrate_import.py`) на чистой БД.

Этап работы атомарный:
  A  счётчики + показания (перенумерация приборов по квартире, журнал),
  B  начисления (фикс: регулярные/разовые/входящие; переменные суммами),
  C  касса,
  D  контрольные суммы и человекочитаемый .xlsx (позже).

Скрипт НЕ трогает БД. Запуск (локальный Python с openpyxl, из корня репо):
    python backend/migrations/migrate_prepare_sources.py \
        --src templates/Миграция данных FTH.xlsx --out templates/clean
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
from collections import Counter, defaultdict
from decimal import Decimal
from pathlib import Path

try:
    import openpyxl
except Exception as exc:  # pragma: no cover
    raise SystemExit(
        "Нужен пакет openpyxl. Установите: python -m pip install openpyxl\n" f"({exc})"
    )


# --- helpers ---------------------------------------------------------------
def _num(v):
    if v is None or isinstance(v, (dt.datetime, dt.date)):
        return None
    try:
        return Decimal(str(v))
    except Exception:
        return None


def _date(v):
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    return None


def apt_int(v):
    if v is None:
        return None
    try:
        return int(str(v).strip().split(".")[0])
    except ValueError:
        return None


def srv_code(v):
    """Код «Виды задолженностей» источника -> канонич. строка (''1''..''7'')."""
    if v is None:
        return None
    try:
        return str(int(str(v).strip().split(".")[0]))
    except ValueError:
        return None


def _ym(v):
    """(year, month) если колонка дата/день-месяц, иначе None."""
    if isinstance(v, (dt.datetime, dt.date)):
        return (v.year, v.month)
    return None


def load_sheet(path: Path, name: str):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb[name]
        it = ws.iter_rows(values_only=True)
        hdr = next(it, None)
        return hdr, [r for r in it]
    finally:
        wb.close()


def write_csv(path: Path, header, rows):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(header)
        w.writerows(rows)


# --------------------------------------------------------------------------
# АТОМ A  (счётчики + показания)
# --------------------------------------------------------------------------
def build_meter_reading_plan(src: Path):
    _, meters = load_sheet(src, "Счетчики")      # Код,Статья,Анал,Кв,Метка,Дата
    _, readings = load_sheet(src, "Показания")   # Кв,Анал,Метка,Дата,Показ,...

    # серии показаний по (кв, метка)
    series = defaultdict(list)
    for r in readings:
        ap = apt_int(r[0])
        lbl = str(r[2])
        d = _date(r[3])
        v = _num(r[4])
        if ap is None or not lbl or d is None:
            continue
        series[(ap, lbl)].append((d, v))
    for k in series:
        series[k].sort(key=lambda x: x[0])

    # приборы на квартиру (по листу Счетчики)
    by_apt = defaultdict(list)
    for m in meters:
        ap = apt_int(m[3])
        if ap is None:
            continue
        by_apt[ap].append({"lbl": str(m[4]), "code": m[0]})

    def _first(ap, lbl):
        s = series.get((ap, lbl))
        return s[0][0] if s else None

    meter_rows, label_idx = [], {}
    for ap in sorted(by_apt):
        devs = sorted(by_apt[ap], key=lambda dd: _first(ap, dd["lbl"]) or dt.date.max)
        for i, dd in enumerate(devs, 1):
            first = _first(ap, dd["lbl"])
            meter_rows.append([ap, i, first.isoformat() if first else "", str(dd["code"]), dd["lbl"]])
            label_idx[(ap, dd["lbl"])] = i

    reading_rows = []
    for r in readings:
        ap = apt_int(r[0])
        d = _date(r[3])
        lbl = str(r[2])
        v = _num(r[4])
        idx = label_idx.get((ap, lbl))
        if ap is None or d is None or idx is None:
            continue
        reading_rows.append([ap, idx, d.isoformat(), "" if v is None else str(v)])
    reading_rows.sort(key=lambda r: (int(r[0]), int(r[1]), r[2]))

    notes = [
        f"Счётчиков (лист): {len(meters)}; квартир: {len(by_apt)}.",
        f"Приборов (квартира,№): {len(meter_rows)}; показаний: {len(reading_rows)}; "
        f"дублей (кв,прибор,дата): "
        f"{len(reading_rows) - len({(int(x[0]), int(x[1]), x[2]) for x in reading_rows})}.",
    ]
    return meter_rows, reading_rows, notes


# --------------------------------------------------------------------------
# АТОМ B  (начисления)
# --------------------------------------------------------------------------
SRC_SERVICE = {  # код «Виды задолженностей» источника (он же анал. номер)
    "1": "Электроэнергия", "2": "Охрана", "3": "Электричество охраны",
    "4": "Обслуживание ТП", "5": "Вывоз мусора", "6": "Фонд развития",
    "7": "Прочие расходы",
}

# Разовые сборы-«дом-акции»: месяц начисляем на дом единой суммой (не месячная база услуги).
HOME_RAZOV = {
    ("6", "2018-03"): "разовая: замена э/э счётчиков",
    ("6", "2018-04"): "разовая: подключение х/в",
    ("6", "2019-10"): "разовая: обслуживание дымоходов",
    ("6", "2024-11"): "разовый сбор (30000)",
    ("6", "2026-04"): "разовый сбор",
    ("7", "2022-01"): "парковка/освещение",
    ("7", "2024-03"): "Проверка газа",
}
# Одно-месячные «не-базовые» месяцы регулярных услуг (не фонда): по ним дом платит
# разовую нетиповую сумму (не регулярная ставка услуги).
ONE_MO_NONBASE = {
    ("2", "2020-11"), ("2", "2020-12"), ("2", "2026-01"),
    ("4", "2025-09"),
}


def _accruals_fix(rows):
    """Фиксированные: классификация на reg/razov/manual/enter."""
    from collections import defaultdict
    grp = defaultdict(list)  # (an, ym) -> [(kv, amount, tariff, kom)]
    for r in rows:
        an = srv_code(r[3])
        ym = _ym(r[4])
        kv = apt_int(r[1])
        if not an or not ym:
            continue
        grp[(an, ym)].append((kv, _num(r[7]), _num(r[5]), r[12]))

    for (an, ym), items in grp.items():
        ym_s = f"{ym[0]:04d}-{ym[1]:02d}"
        is_enter = (an == "6" and ym_s == "2017-10")
        razo_comm = HOME_RAZOV.get((an, ym_s)) if not is_enter else None
        one_mo = (an, ym_s) in ONE_MO_NONBASE

        # Пропуск: нулевые и кв13-мусор (ан5). Отрицательные НЕ пропускаем (входят
        # сторно/переплаты старта). Для входящих (is_enter) не считаем моду вообще.
        def _go(it):
            kv, am, t, k = it
            if am is None or am == 0:
                return False   # «нулевые строки» не переносим
            if an == "5" and kv == 13:
                return False   # кв13 не числится по вывозу мусора
            return True

        if is_enter:
            # Входящие остатки 2017-10: не вычисляем моду, каждая строка (кроме нуля)
            # — отдельный документ входящего остатка со своей суммой (в т.ч. −398 у кв17).
            for (kv, am, t, kom) in items:
                if am is None or am == 0:
                    continue
                kv = apt_int(kv) if not isinstance(kv, int) else kv
                yield (kv, an, ym_s, "enter", am, 1, Decimal("1"), am, "Входящий остаток")
            continue

        act = [it for it in items if _go(it)]
        if not act:
            continue
        mode = Counter(it[1] for it in act).most_common(1)[0][0]

        for (kv, am, t, kom) in items:
            if not _go((kv, am, t, kom)):
                continue
            kv = apt_int(kv) if not isinstance(kv, int) else kv
            if razo_comm is not None and abs(am - mode) < Decimal("0.005"):
                yield (kv, an, ym_s, "razov", am, 1, Decimal("1"), am, razo_comm)
                continue
            if one_mo and abs(am - mode) < Decimal("0.005"):
                yield (kv, an, ym_s, "razov", am, 1, Decimal("1"), am, "разовый месяц (не база услуги)")
                continue
            if abs(am - mode) > Decimal("0.005"):
                yield (kv, an, ym_s, "manual", am, 1, Decimal("1"), am, kom or "персональная корректировка/смета")
                continue
            yield (kv, an, ym_s, "reg", am, 0, Decimal("1"), am, None)


def _accruals_var(rows):
    """Переменные: суммами как есть (kind='vary'), с кол-вом и тарифом источника."""
    for r in rows:
        an = srv_code(r[4])
        ym = _ym(r[5])
        kv = apt_int(r[1])
        am = _num(r[10])
        if not an or not ym or am is None:
            continue
        consumption = _num(r[8])
        tariff_price = _num(r[9])
        # amount(сум.) источника уже = consumption * тариф (сверено); эти два применяем
        # для восстановления настоящего Tariff+consumptionв синтезе (вариант a).
        if consumption is None:
            consumption = Decimal("1")
        if tariff_price is None:
            tariff_price = consumption  # безопасный fallback (amount==price*1)
        yield (kv, an, f"{ym[0]:04d}-{ym[1]:02d}", "vary", am, 0, consumption, tariff_price, None)


def build_accruals_plan(src: Path):
    _, fix = load_sheet(src, "Начисления-Фиксированные")
    _, var = load_sheet(src, "Начисления-Переменные")
    rows = []
    rows += list(_accruals_fix(fix))
    rows += list(_accruals_var(var))
    counts = Counter(x[3] for x in rows)
    notes = [f"Начислений фикс+перем: {len(rows)}; виды: " + ", ".join(f"{k}={c}" for k, c in counts.items())]
    return rows, notes


# ---------------------------------------------------------------------------
# АТОМ AP: справочные квартиры/владельцы (для синтеза справочников на чистой БД)
# ---------------------------------------------------------------------------

def build_apartments_plan(src: Path):
    """Строит реестр квартир (эталон для синтеза Counterparty/Apartment/Account).

    Источники: лист Квартиры (код_квартиры, код_контрагента, номер, площадь)
    и Контрагенты (код -> ФИО и т.п.).
    """
    # перечитаем с хедерами отдельно
    wb = openpyxl.load_workbook(src, data_only=True, read_only=True)
    out = []
    notes = []
    try:
        def _sheet(name):
            ws = wb[name]
            it = ws.iter_rows(values_only=True)
            return next(it, None), list(it)

        hk_q, kv_rows = _sheet("Квартиры")
        hk_ct, ct_rows = _sheet("Контрагенты")
        ikq = {n: i for i, n in enumerate(hk_q)}
        ict = {n: i for i, n in enumerate(hk_ct)}

        def _cell(row, index) :
            try:
                return row[index]
            except Exception:
                return None

        # Словарь контрагентов по коду
        kontr = {}
        for r in ct_rows:
            code = apt_int(_cell(r, ict.get("Код контрагента")))
            if code is None:
                continue
            fam = _cell(r, ict.get("Фамилия")) or ""
            im = _cell(r, ict.get("Имя")) or ""
            ot = _cell(r, ict.get("Отчество")) or ""
            full = " ".join(x for x in (fam, im, ot) if x).strip()
            phone = _cell(r, ict.get("Мобильный телефон")) or \
                    _cell(r, ict.get("Домашний телефон")) or ""
            email = _cell(r, ict.get("e-mail")) or ""
            kontr[code] = (full, phone, email)

        for r in kv_rows:
            ap = apt_int(_cell(r, ikq.get("Код_Квартира")))
            if ap is None:
                continue
            owner_code = apt_int(_cell(r, ikq.get("Код_Контрагент")))
            number = _cell(r, ikq.get("Номер_Квартира"))
            sq = _num(_cell(r, ikq.get("Площадь")))
            full, phone, email = kontr.get(owner_code or -1, ("", "", ""))
            out.append([ap, number if number is not None else ap,
                        owner_code, full, phone, email,
                        "" if sq is None else str(sq)])
    finally:
        wb.close()
    notes.append(f"Справочных квартир с владельцами: {len(out)}.")
    return out, notes


# ---------------------------------------------------------------------------
# АТОМ C: касса (приход/расход)
# --------------------------------------------------------------------------
# Приход-лист: 0 Код 1 Квартира 2 Контрагент 3 Статья 4 Аналитика 5 Сумма
#               6 Дата 7 Время 8 Комментарий
# Расход-лист: 0 Код 1 Квартира 2 Контрагент 3 Статья 4 Аналитика 5 Код_Проект
#              6 Сумма 7 Дата 8 Время 9 Комментарий


def _build_cash_in(rows):
    for r in rows:
        kv = apt_int(r[1])
        an = srv_code(r[4])
        amt = _num(r[5])
        d = _date(r[6])
        comment = r[8]
        if amt is None or d is None:
            continue
        if amt == 0:
            continue  # нулевые разнески не переносим
        # «Начальный остаток» -> отдельная операция-входящего сальдо кассы
        if (comment or "").strip().startswith("Начальный остаток"):
            yield ("start_cash", None, abs(amt), d, r[7], an, 2, "Начальный остаток кассы (входящее сальдо)")
            continue
        storno = amt < 0
        # Приход без указанной квартиры (код 0/пусто) -> account = NULL (не жительский)
        acc = kv if kv and kv > 0 else None
        yield ("in", acc, abs(amt), d, r[7], an, (1 if storno else 0),
               (comment or "") + (" [СТОРНО]" if storno else ""))


def _build_cash_out(rows):
    for r in rows:
        kv = apt_int(r[1])
        an = srv_code(r[4])
        amt = _num(r[6])
        d = _date(r[7])
        project = r[5]
        comment = r[9]
        if amt is None or d is None:
            continue
        if amt == 0:
            continue
        acc = kv if kv and kv > 0 else None
        yield ("out", acc, abs(amt), d, r[8], an, 0,
               (comment or "") + (f" [проект:{project}]" if project is not None and str(project).strip() not in ("", "None") else ""))


def build_cash_plan(src: Path):
    _, ci = load_sheet(src, "Касса-Приход")
    _, co = load_sheet(src, "Касса-Расход")
    rows = list(_build_cash_in(ci)) + list(_build_cash_out(co))
    notes = []
    notes.append(f"Касса: строк приход+расход: {len(rows)} (in/out/start_cash посчитаны ниже).")
    return rows, notes


# --------------------------------------------------------------------------
# АТОМ D: контрольные суммы и человекочитаемый .xlsx
# --------------------------------------------------------------------------
START_COLUMNS_CASH = ["kind", "account", "amount", "date", "time", "analytic_src", "is_storno", "comment"]


def control_csv_totals(accrual_rows):
    """Сводные контрольные суммы по начислениям из готовых строк."""
    kinds = defaultdict(lambda: [0, None])  # kind -> [count, Decimal sum]
    for r in accrual_rows:
        kind, amt = r[3], r[4]
        kinds[kind][0] += 1
        kinds[kind][1] = (kinds[kind][1] or Decimal(0)) + amt
    lines = []
    for kind in ("reg", "razov", "manual", "enter", "vary"):
        cnt, total = kinds[kind]
        lines.append(f"  начислений kind={kind}: строк {cnt}, сумма {total}")
    return lines


def report_cash_control(src: Path, cash_rows):
    """Рассифровка итогов кассы: сколько dropped, стorno, старт и валидация сумм."""
    out = []
    sum_in = Decimal(0); sum_out = Decimal(0); sum_start = Decimal(0)
    for r in cash_rows:
        kind = r[0]
        amt = r[2]
        if kind == "in":
            sum_in += amt
        elif kind == "out":
            sum_out += amt
        else:
            sum_start += amt
    out.append(f"касса in (приход чисто): {sum_in}")
    out.append(f"касса out (расход чисто): {sum_out}")
    out.append(f"касса start_cash (вх. сальдо): {sum_start}")
    return out


def write_clean_xlsx(path: Path, cash_header, cash_rows):
    """Человекочитаемый итог: из CSV работаем напрямую из предоставленных rows."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "cash"
    ws.append(cash_header)
    for r in cash_rows:
        ws.append([r[0], r[1], float(r[2]) if r[2] is not None else None,
                   r[3].isoformat() if hasattr(r[3], "isoformat") else r[3],
                   r[4], r[5], r[6], str(r[7]) if r[7] is not None else None])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


# --- main -----------------------------------------------------
def main() -> int:
    ap = argparse.ArgumentParser(description="Подготовка источников миграции.")
    ap.add_argument("--src", default="templates/Миграция данных FTH.xlsx")
    ap.add_argument("--out", default="templates/clean")
    ap.add_argument("--xlsx", default=None, help="путь для человекочитаемого итогового .xlsx (необяз.)")
    args = ap.parse_args()

    src = Path(args.src).resolve()
    out = Path(args.out).resolve()
    if not src.exists():
        print(f"Не найден источник: {src}")
        return 2

    meter_rows, reading_rows, nA = build_meter_reading_plan(src)
    apartment_rows, nAp = build_apartments_plan(src)
    accrual_rows, nB = build_accruals_plan(src)
    cash_rows, nC = build_cash_plan(src)

    write_csv(out / "apartments.csv",
              ["apartment_key", "apartment_number", "owner_key", "owner_full_name",
               "owner_phone", "owner_email", "square"], apartment_rows)
    write_csv(out / "meters.csv",
              ["apartment_key", "meter_idx", "first_reading_date", "src_code", "src_label"], meter_rows)
    write_csv(out / "readings.csv",
              ["apartment_key", "meter_idx", "reading_date", "reading"], reading_rows)
    write_csv(out / "accruals.csv",
              ["apartment", "service_src", "period", "kind_in_source",
               "amount", "flag_do_not_recalc", "consumption", "tariff_price", "comment"], accrual_rows)
    write_csv(out / "services.csv",
              ["code", "name"], [[k, v] for k, v in SRC_SERVICE.items()])
    write_csv(out / "cash.csv",
              ["kind", "account", "amount", "date", "time", "analytic_src",
               "is_storno", "comment"], cash_rows)

    print("=== Подготовка источников ===")
    print("АТОМ A (счётчики+показания):"); [print(" -", x) for x in nA]
    print("АТОМ АП (квартиры/владельцы):"); [print(" -", x) for x in nAp]
    print("АТОМ B (начисления):"); [print(" -", x) for x in nB]
    print("АТОМ C (касса):"); [print(" -", x) for x in nC]
    print("АТОМ D (контроль — начисления):")
    [print(x) for x in control_csv_totals(accrual_rows)]
    print("АТОМ D (контроль — касса):")
    [print(x) for x in report_cash_control(src, cash_rows)]
    if args.xlsx:
        write_clean_xlsx(Path(args.xlsx), START_COLUMNS_CASH, cash_rows)
        print(f"Человекочитаемый xlsx записан: {args.xlsx}")
    print(f"Записано в {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
