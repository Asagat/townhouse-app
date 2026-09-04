# backend/migrations/reimport/reader.py
"""Чтение листов «Касса-Приход»/«Касса-Расход» из templates/Миграция данных FTH.xlsx
в нормализованный список записей БЕЗ слияния: по приходам каждый плательщик — свой
документ; по расходам каждый расход — отдельная запись.

Выдаёт только структуру (dry-run/отчёт); запись в БД — следующий контрольный шаг.
Признаки «входящего остатка»: комментарий «начальн…/остаток/корректировк/входящ/сторно»
или отрицательная сумма → type 'opening' (не доход/расход).
"""

from __future__ import annotations

from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date
import sys
from pathlib import Path
from typing import Any, Iterable


@dataclass
class CashRec:
    flow: str            # 'in' | 'out'
    side: str            # 'income' | 'expense' | 'opening'
    is_opening: bool
    amount: float
    day: date | None
    apartment: int | None
    counterparty_code: int | None
    comment: str = ""
    time: str = ""


def _num(v) -> float | None:
    try:
        if v is None:
            return None
        return float(v)
    except Exception:
        return None


def _icode(v) -> int | None:
    if v is None:
        return None
    try:
        return int(float(v))
    except Exception:
        return None


def _opening(comment: str, amount: float) -> bool:
    low = (comment or "").lower()
    if any(k in low for k in ("начальн", "остаток", "корректировк", "входящ", "сторно", "возврат ош")):
        return True
    return amount < 0


def _day(v) -> date | None:
    if v is None:
        return None
    try:
        return v.date() if hasattr(v, "date") else v
    except Exception:
        return None


def _sheet_map(wb, name):
    ws = wb[name]
    headers = [c.value for c in next(ws.iter_rows(max_row=1))]
    rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if r and any(x is not None for x in r)]
    return ws, headers, rows


def parse(path: str | Path) -> list[CashRec]:
    import openpyxl
    wb = openpyxl.load_workbook(str(path), data_only=True)
    recs: list[CashRec] = []

    # Касса-Приход
    ws, headers, rows = _sheet_map(wb, "Касса-Приход")
    ix = {h: i for i, h in enumerate(headers) if h is not None}
    amt = ix.get("Сумма")
    for r in rows:
        a = _num(r[amt]) if amt is not None and r[amt] is not None else None
        if a is None:
            continue
        com = r[ix.get("Комментарий")] or ""
        recs.append(CashRec(
            flow="in",
            side="opening" if _opening(com, a) else "income",
            is_opening=_opening(com, a), amount=a,
            day=_day(r[ix["Дата"]] if "Дата" in ix else None),
            apartment=_icode(r[ix["Код_Квартира"]]) if "Код_Квартира" in ix else None,
            counterparty_code=_icode(r[ix["Код_Контрагент"]]) if "Код_Контрагент" in ix else None,
            comment=com,
            time=str(r[ix["Время"]] or "") if "Время" in ix else "",
        ))

    # Касса-Расход
    ws2, h2, rows2 = _sheet_map(wb, "Касса-Расход")
    ix2 = {h: i for i, h in enumerate(h2) if h is not None}
    amt2 = ix2.get("Сумма")
    for r in rows2:
        a = _num(r[amt2]) if amt2 is not None and r[amt2] is not None else None
        if a is None:
            continue
        com = r[ix2.get("Комментарий")] or ""
        recs.append(CashRec(
            flow="out",
            side="opening" if _opening(com, a) else "expense",
            is_opening=_opening(com, a), amount=a,
            day=_day(r[ix2["Дата"]] if "Дата" in ix2 else None),
            apartment=_icode(r[ix2["Квартира"]]) if "Квартира" in ix2 else None,
            counterparty_code=_icode(r[ix2["Контрагент"]]) if "Контрагент" in ix2 else None,
            comment=com,
            time=str(r[ix2["Время"]] or "") if "Время" in ix2 else "",
        ))
    return recs


def summarize(recs: list[CashRec]) -> dict:
    s = {"count": len(recs), "by_side": Counter(r.side for r in recs),
         "by_flow": Counter(r.flow for r in recs), "sum": defaultdict(float)}
    s["years"] = Counter()
    s["no_apt"] = Counter(r.side for r in recs if r.apartment is None)
    for r in recs:
        s["sum"][r.side] += r.amount
        if r.day:
            s["years"][r.day.year] += 1
    return dict(s)


def main() -> None:
    if len(sys.argv) < 2:
        print("Укажите путь к xlsx")
        return
    xlsx = sys.argv[1]
    recs = parse(xlsx)
    i = summarize(recs)
    print("rows:", i["count"])
    print("by_side:", dict(i["by_side"]))
    print("by_flow:", dict(i["by_flow"]))
    print("sum:", {k: round(v, 2) for k, v in i["sum"].items()})
    print("years:", dict(sorted(i["years"].items())))
    print("no_apartment rows:", dict(i["no_apt"]))


if __name__ == "__main__":
    main()
