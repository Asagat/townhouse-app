# backend/migrations/reimport/build_cash.py
"""Milestone B-2: черновой залив кассы в staging по нормализованным строкам reader.

Только на townhouse_stage (защита через env STAGE_DB). Пишет transactions + cash_register
core-SQL (без ORM-событий), назначает article_id/contractor_id по маппингам, проставляет
doc_no по хронологии. До этого очищает transactions/cash_register/accounts_register.

Отчёт: суммы/кол-ва, список "битых" ссылок (нет справочника/квартиры).
"""

import os
import re
import sys
from decimal import Decimal
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]  # backend
sys.path.insert(0, str(ROOT))

import database  # noqa: E402
from migrations.reimport.reader import parse, CashRec  # noqa: E402
from models import recalculate_cash_balance  # noqa: E402
from sqlalchemy import text  # noqa: E402

# правила расходов (зеркало migrate_expand_article_2026.RULES)
_RULES = [
    ("Электроэнергия", ["эл.энерг", "эл. энерги", "электроэнерг", "электричеств", "за свет", "освещ"]),
    ("Обслуживание инженерных систем", ["подключен холодн", "холодной воды", "кабел", "трансформатор", "тп", "канализац", "дымоход", "счетчик", "счётчик", "электрику", "электр.", "фонар", "аварийн", "повреждений", "модели для", "трубы"]),
    ("Водоснабжение", ["вод", "поливн"]),
    ("Заработная плата персонала", ["охране", "охрану", "зп", "зарплат", "заработ", "окончательный расчет", "оразу за работу", "выдач", "ремонт охраны", "ремонт дома охраны", "за обогревател для ох", "эл.плита для охран"]),
    ("Благоустройство территории", ["асфальт", "арык", "арычн", "лотк", "грунт", "чернозем", "детск", "дет.площадк", "воркаут", "баскетб", "стойку", "кольцо", "сетка", "парковк", "покрыт", "песочниц", "песок", "озелен", "удобрен", "клён", "клен", "газон", "решетк", "шлагбаум", "снегоубороч", "салют", "качел", "гирлянд", "украшен", "бесед", "желез", "кровл", "навес", "выключател", "3д панел", "доставк", "мягк"]),
    ("Вывоз мусора и утилизация", ["мусор", "утил", "вывоз грунт"]),
    ("Безопасность и проверки", ["видеонаблюден", "камеры", "чс", "пожар", "безопасн", "сигнализ", "топосъем", "экспертное"]),
    ("Материалы и инвентарь", ["бачк", "соль", "материал", "бетон", "арматур", "инвентар", "лопат", "тачк", "перчатк", "оборуд", "запчаст", "шкаф", "краск", "хлорк", "ножниц", "заклеп", "кран", "шланг", "лент", "метл", "совок", "пистолет", "смесител", "сигнальная", "на материлы", "хозтовар"]),
    ("Возвраты жителям", ["возврат", "вернул", "перерасчет", "взаиморасчёт", "расходам на квартиру"]),
]
_FALLBACK = "Прочие расходы"


def _expense_article(comment: str) -> str:
    low = (comment or "").lower()
    for name, keys in _RULES:
        for k in keys:
            if k in low:
                return name
    return _FALLBACK


def _income_article(comment: str) -> str:
    low = (comment or "").lower()
    if any(k in low for k in ("возврат", "перерасчет", "ошиб")):
        return "Возвраты от контрагентов"
    # обычные платежи и фонд развития
    return "Поступления от жителей"


def _norm(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip().lower())


def main() -> int:
    dburl = database.SQLALCHEMY_DATABASE_URL
    if "townhouse_stage" not in dburl:
        print("Отказ: строительство разрешено только на townhouse_stage.")
        return 3

    xlsx = sys.argv[1] if len(sys.argv) > 1 else "Миграция данных FTH.xlsx"
    recs: list[CashRec] = parse(xlsx)
    recs.sort(key=lambda r: (r.day or __import__("datetime").date.min, r.flow, r.amount))

    db = database.SessionLocal()
    try:
        # справочники-статьи (kind маппится по имени)
        def art_id(name: str) -> int:
            row = db.execute(text("SELECT id FROM analytic_articles WHERE name=:n LIMIT 1"),
                             {"n": name}).first()
            if row:
                return row[0]
            # kind income/expense/opening
            kind = {"Поступления от жителей": "income", "Возвраты от контрагентов": "income",
                    "Прочие доходы": "income", "Входящий остаток": "opening"}.get(name, "expense")
            rid = db.execute(text(
                "INSERT INTO analytic_articles (name, kind, is_active) VALUES (:n,:k,true) RETURNING id"
            ), {"n": name, "k": kind}).scalar()
            return rid

        opening_art = art_id("Входящий остаток")

        # квартиры/accounts
        apt_owner = {}
        acc_by_apt = {}
        for a, acc, own in db.execute(text(
            "SELECT ap.apartment_number, acc.id, ap.owner_id FROM apartments ap "
            "LEFT JOIN accounts acc ON acc.apartment_id = ap.id "
            "LEFT JOIN counterparties c ON c.id=ap.owner_id"
        )).fetchall():
            apt_owner[int(a)] = own
            if acc:
                acc_by_apt[int(a)] = acc
        fallback_c = db.execute(text("SELECT id FROM counterparties WHERE full_name LIKE 'Сагатбекова Людмила%'")).scalar()

        # Таблица кодов контрагентов по имени листа (файл: там же, где xlsx)
        import openpyxl
        wb = openpyxl.load_workbook(xlsx, data_only=True)
        cws = wb["Контрагенты"]
        hd = [c.value for c in next(cws.iter_rows(max_row=1))]
        ix = {h: i for i, h in enumerate(hd) if h}
        name_of_code = {}
        for r in cws.iter_rows(min_row=2, values_only=True):
            if r is None or r[ix["Код контрагента"]] is None:
                continue
            code = int(float(r[ix["Код контрагента"]]))
            nm = " ".join(str(r[ix[k]] or "").strip() for k in ("Фамилия", "Имя", "Отчество") if k in ix)
            name_of_code[code] = _norm(nm)
        cpmap = {}
        for cid, nm in db.execute(text("SELECT id, full_name FROM counterparties")).fetchall():
            cpmap.setdefault(_norm(nm), cid)

        # очистка истории кассы (только stage)
        db.execute(text("DELETE FROM cash_register"))
        db.execute(text("DELETE FROM accounts_register"))
        db.execute(text("DELETE FROM transactions"))

        tx_ins = text("INSERT INTO transactions (transaction_date, created_at, account_id, cash_point_id, "
                      "article_id, contractor_id, transaction_type, amount, notes, title, doc_no) "
                      "VALUES (:d,:c,:acc,:cp,:art,:ctr,:ttype,:amt,:notes,:title,:no) RETURNING id")
        cr_ins = text("INSERT INTO cash_register (operation_date, account_id, transaction_id, contractor_id, "
                      "income, expense, balance_after) VALUES (:d,:acc,:tid,:ctr,:inc,:exp,0)")

        cp_id = db.execute(text("SELECT id FROM cash_points LIMIT 1")).scalar()
        cash_point_name = db.execute(text("SELECT name FROM cash_points LIMIT 1")).scalar() or "Касса"

        n = 0
        bad_refs = []
        for seq, r in enumerate(recs, start=1):
            account_id = acc_by_apt.get(r.apartment) if r.apartment is not None else None

            # статья
            if r.side == "opening":
                art = opening_art
            elif r.flow == "out":
                art = art_id(_expense_article(r.comment))
            else:
                art = art_id(_income_article(r.comment))

            # контрагент: приход -> владелец квартиры, иначе код->имя->справочник или Людмила
            if r.flow == "in" and r.apartment is not None and apt_owner.get(r.apartment):
                ctr = apt_owner[r.apartment]
            elif r.counterparty_code is not None:
                nm = name_of_code.get(r.counterparty_code)
                ctr = cpmap.get(nm) if nm else None
            else:
                ctr = None
            if not ctr:
                ctr = fallback_c
                if ctr is None:
                    bad_refs.append(("counterparty", r.apartment, r.counterparty_code, r.comment))

            if account_id is None and r.apartment is not None:
                bad_refs.append(("account/apartment", r.apartment, None, None))
            if art is None:
                bad_refs.append(("article", None, None, r.comment))

            d = r.day or __import__("datetime").date.min
            is_in = r.flow == "in"
            ttype = "in_cash" if is_in else "out_cash"
            inc = Decimal(r.amount) if is_in else Decimal("0")
            exp = Decimal(r.amount) if not is_in else Decimal("0")

            title = f"{'Приход в кассу' if is_in else 'Расход из кассы'} №{seq} от {d:%d.%m.%Y}"
            tid = db.execute(tx_ins, {
                "d": d, "c": d, "acc": account_id, "cp": cp_id, "art": art, "ctr": ctr,
                "ttype": ttype, "amt": r.amount, "notes": r.comment, "title": title, "no": seq,
            }).scalar()
            db.execute(cr_ins, {"d": d, "acc": account_id, "tid": tid, "ctr": ctr,
                                "inc": inc if is_in else Decimal("0"), "exp": Decimal("0") if is_in else exp})
            n += 1
        # Подсчёт балансов по счёту с нуля (пересчёт оконной функции)
        acc_ids = [x[0] for x in db.execute(text(
            "SELECT DISTINCT account_id FROM cash_register WHERE account_id IS NOT NULL")).fetchall()]
        for aid in acc_ids:
            recalculate_cash_balance(db, aid)
        db.commit()

        print(f"Записано cash-документов: {n}")
        print("битых ссылок:", len(bad_refs), "— примеры:", bad_refs[:8])
        sums = db.execute(text("""
          SELECT a.kind, sum(t.amount) FROM transactions t JOIN analytic_articles a ON a.id=t.article_id
          GROUP BY a.kind ORDER BY a.kind
        """)).fetchall()
        for kind, s in sums:
            print(f"   статья-kind {kind}: {float(s or 0):,.2f}")
        return 0
    except Exception as exc:  # noqa: BLE001
        db.rollback()
        print("ОШИБКА:", exc)
        return 1
    finally:
        db.close()


if __name__ == "__main__":
    raise SystemExit(main())
