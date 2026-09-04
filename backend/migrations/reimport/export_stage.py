# backend/migrations/reimport/export_stage.py
"""Полная выгрузка townhouse_stage в один .xlsx (лист на каждую таблицу).

Запуск от каталога backend с DATABASE_URL на townhouse_stage:
    python migrations/reimport/export_stage.py [out.xlsx]
"""

import os
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

import database  # noqa: E402
from openpyxl import Workbook
from sqlalchemy import text  # noqa: E402


def main() -> None:
    db = database.SessionLocal()
    out = Path(sys.argv[1] if len(sys.argv) > 1 else "stage_dump.xlsx")
    wb = Workbook()
    wb.remove(wb.active)
    try:
        tables = [r[0] for r in db.execute(text(
            "SELECT table_name FROM information_schema.tables "
            "WHERE table_schema='public' AND table_name NOT LIKE 'alembic_version' "
            "ORDER BY table_name"
        )).fetchall()]
        summary = wb.create_sheet("_список")
        summary.append(["таблица", "строк"])
        total_rows = 0
        for t in tables:
            cols = [c[0] for c in db.execute(text(
                "SELECT column_name FROM information_schema.columns WHERE table_name=:t "
                "ORDER BY ordinal_position"), {"t": t}).fetchall()]
            ws = wb.create_sheet(t[:31])
            ws.append([str(c) for c in cols])
            rows = db.execute(text(f'SELECT * FROM "{t}"')).all()
            for r in rows:
                vals = ["" if c is None else str(c) for c in tuple(r)]
                ws.append(vals)
            summary.append([t, len(rows)])
            total_rows += len(rows)
        wb.save(str(out))
        print("Таблиц:", len(tables), "строк всего:", total_rows)
        print("Файл: /app/%s (хост: backend/%s)" % (out.name, out.name))
    except Exception as exc:  # noqa: BLE001
        print("Ошибка:", exc)
    finally:
        db.close()


if __name__ == "__main__":
    main()
